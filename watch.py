"""
Simple File Processor - Monitors upload/ folder

Drop Excel files (.xlsm, .xlsx) into the upload/ folder.
They are automatically processed and moved to archive/.
Output appears in the root folder as filename.xlsx

Just run: python watch.py
And leave it running!
"""

import os
import sys
import time
import logging
import shutil
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
import pandas as pd
import re
from difflib import get_close_matches

# ===== SETUP =====
UPLOAD_FOLDER = "upload"
ARCHIVE_FOLDER = "archive"
OUTPUT_FOLDER = "."
MARGIN_PERCENT = 2.75
SUPPORTED_EXTENSIONS = ('.xlsm', '.xlsx', '.xls')

# Ensure folders exist
Path(UPLOAD_FOLDER).mkdir(exist_ok=True)
Path(ARCHIVE_FOLDER).mkdir(exist_ok=True)

# Setup logging
log_folder = "logs"
Path(log_folder).mkdir(exist_ok=True)
log_file = os.path.join(log_folder, f"etl_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log")

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_file),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# ===== HELPER FUNCTIONS =====

def normalize(col):
    if col is None:
        return ""
    col = str(col)
    col = col.strip().lower()
    col = re.sub(r"\s+", " ", col)
    return col


def find_best_match(col_names, search_terms, verbose=False):
    """Find best matching column"""
    normalized = {col: normalize(col) for col in col_names}

    # exact match
    for col, n in normalized.items():
        if all(t in n for t in search_terms):
            if verbose:
                logger.debug(f"Direct match: {col}")
            return col

    # partial match
    for col, n in normalized.items():
        if any(t in n for t in search_terms):
            if verbose:
                logger.debug(f"Partial match: {col}")
            return col

    # token scoring
    best = None
    best_score = 0
    for col in col_names:
        tokens = normalize(col).split()
        score = sum(1 for t in search_terms if t in tokens)
        if score > best_score:
            best_score = score
            best = col
    if best_score > 0:
        return best

    # difflib
    candidates = [str(c) for c in col_names if c is not None and str(c) != 'nan']
    closers = get_close_matches(" ".join(search_terms), candidates, n=1, cutoff=0.6)
    if closers:
        return closers[0]

    return None


def detect_header_row(df_obj, tokens, scan_rows=15, min_matches=4):
    """Find header row"""
    scan_rows = min(scan_rows, max(0, df_obj.shape[0]))
    tokens = [t.lower() for t in tokens]
    for i in range(scan_rows):
        row_vals = [normalize(v) for v in df_obj.iloc[i].tolist()]
        count = sum(1 for t in tokens if any(t in (str_cell or "") for str_cell in row_vals))
        if count >= min_matches:
            return i
    return None


def find_value_near_key(ws, key_tokens):
    """Extract value from Cover sheet"""
    key_tokens = [k.lower() for k in key_tokens]
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                val_norm = normalize(cell.value)
                for token in key_tokens:
                    if token in val_norm:
                        row_vals = [c for c in row]
                        try:
                            idx = row_vals.index(cell)
                        except ValueError:
                            idx = None
                        if idx is not None:
                            if idx + 1 < len(row_vals):
                                next_cell = row_vals[idx + 1]
                                if next_cell.value and isinstance(next_cell.value, str):
                                    if next_cell.value.strip():
                                        return next_cell.value.strip()
                            if cell.row + 1 <= ws.max_row:
                                below_cell = ws.cell(row=cell.row + 1, column=cell.column)
                                if below_cell.value and isinstance(below_cell.value, str):
                                    if below_cell.value.strip():
                                        return below_cell.value.strip()
    return None


def extract_cover_and_customer(ws):
    """Extract metadata from Cover sheet"""
    cover_keys = [
        (["quotation", "#"], "Quotation #"),
        (["qdr", "#"], "QDR #"),
        (["spr", "#"], "SPR #"),
        (["opportunity", "#"], "Opportunity #"),
        (["quote", "name"], "Quote Name"),
        (["quotation", "date"], "Quotation Date"),
        (["valid", "until"], "Valid Until"),
    ]
    customer_keys = [
        (["contact", "name"], "Contact Name"),
        (["company"], "Company"),
        (["address"], "Address"),
        (["city", "state", "zip"], "City, State ZIP"),
        (["country"], "Country"),
        (["phone"], "Phone Number"),
        (["e-mail", "email"], "E-mail"),
    ]
    
    cover_details = {}
    for tokens, label in cover_keys:
        val = find_value_near_key(ws, tokens)
        cover_details[label] = val if val else "N/A"
    
    customer_details = {}
    for tokens, label in customer_keys:
        val = find_value_near_key(ws, tokens)
        customer_details[label] = val if val else "N/A"
    
    return cover_details, customer_details


def process_file(input_file):
    """Process Excel file"""
    logger.info(f"Processing: {input_file}")
    
    try:
        # Get base name for output
        base_name = os.path.splitext(os.path.basename(input_file))[0]
        output_file = os.path.join(OUTPUT_FOLDER, f"{base_name}.xlsx")
        
        # Detect visible sheets
        wb_detect = load_workbook(input_file)
        visible_sheets = [name for name in wb_detect.sheetnames if wb_detect[name].sheet_state == "visible"]
        wb_detect.close()
        
        logger.info(f"Visible sheets: {visible_sheets}")
        
        # Read sheets
        df_dict = pd.read_excel(input_file, sheet_name=visible_sheets)
        if not isinstance(df_dict, dict):
            df_dict = {visible_sheets[0]: df_dict}
        
        transformed_rows = {}
        processed_sheets = []
        
        # Process each sheet
        for sheet_name, df in df_dict.items():
            logger.info(f"Processing sheet: {sheet_name}")
            
            tokens = ["model", "qty", "quantity", "net", "price"]
            header_row_index = detect_header_row(df, tokens, scan_rows=15, min_matches=4)
            
            if header_row_index is not None and header_row_index > 0:
                new_cols = df.iloc[header_row_index].tolist()
                df = df.iloc[header_row_index + 1:].copy().reset_index(drop=True)
                df.columns = [str(c).strip() if c and not str(c).startswith("Unnamed") else f"col_{i}" for i, c in enumerate(new_cols)]
            
            col_names = [str(c) for c in df.columns]
            
            # Map columns
            model_col = find_best_match(col_names, ["model"])
            qty_col = find_best_match(col_names, ["qty", "quantity"])
            net_price_col = find_best_match(col_names, ["net", "price"])
            
            if not (model_col and qty_col and net_price_col):
                logger.info(f"Skipped {sheet_name} - missing columns")
                continue
            
            # Transform rows
            rows = []
            for _, row in df.iterrows():
                model = row.get(model_col, "")
                orig_model = str(model).strip()
                
                if not orig_model or pd.isna(row.get(model_col)):
                    continue
                
                try:
                    qty = float(row[qty_col]) if pd.notna(row[qty_col]) else 0
                except:
                    qty = 0
                
                try:
                    price = float(row[net_price_col]) if pd.notna(row[net_price_col]) else 0.0
                except:
                    price = 0.0
                
                if isinstance(qty, (int, float)) and qty == 0:
                    continue
                
                model_upper = orig_model.upper()
                if model_upper != "TARIFF":
                    denom = 1 - (MARGIN_PERCENT / 100)
                    nprice = price / denom if denom != 0 else price
                    if nprice == 0:
                        nprice = 0.01
                else:
                    nprice = price
                
                formatted = f"{orig_model} {qty} {nprice:0.2f},*,*,*,{price:0.2f}"
                rows.append(formatted)
            
            logger.info(f"Built {len(rows)} rows from {sheet_name}")
            transformed_rows[sheet_name] = rows
            processed_sheets.append(sheet_name)
        
        # Create output
        if transformed_rows:
            out_wb = Workbook()
            default = out_wb.active
            out_wb.remove(default)
            
            consolidated_ws = out_wb.create_sheet(title="Consolidated")
            row_cursor = 1
            
            # Add Cover details
            try:
                wb2 = load_workbook(input_file, data_only=True, keep_vba=True)
                ws_cover = None
                for name in wb2.sheetnames:
                    if name.lower().strip() == 'cover':
                        ws_cover = wb2[name]
                        break
                
                if ws_cover is not None:
                    cover_details, customer_details = extract_cover_and_customer(ws_cover)
                    consolidated_ws.cell(row=row_cursor, column=1, value="Cover").font = Font(size=16, bold=True)
                    row_cursor += 1
                    consolidated_ws.cell(row=row_cursor, column=1, value="Cover Details:")
                    row_cursor += 1
                    for k, v in cover_details.items():
                        consolidated_ws.cell(row=row_cursor, column=1, value=k)
                        consolidated_ws.cell(row=row_cursor, column=2, value=v)
                        row_cursor += 1
                    row_cursor += 1
                    consolidated_ws.cell(row=row_cursor, column=1, value="Customer Details:")
                    row_cursor += 1
                    for k, v in customer_details.items():
                        consolidated_ws.cell(row=row_cursor, column=1, value=k)
                        consolidated_ws.cell(row=row_cursor, column=2, value=v)
                        row_cursor += 1
                    row_cursor += 2
                wb2.close()
            except Exception as e:
                logger.error(f"Error extracting Cover: {e}")
            
            # Add sheet data
            try:
                for sheet_name, rows in transformed_rows.items():
                    consolidated_ws.cell(row=row_cursor, column=1, value=sheet_name).font = Font(size=14, bold=True)
                    row_cursor += 1
                    for s in rows:
                        consolidated_ws.cell(row=row_cursor, column=1, value=s)
                        row_cursor += 1
                    row_cursor += 1
            except Exception as e:
                logger.error(f"Error writing data: {e}")
            
            # Save
            try:
                out_wb.save(output_file)
                logger.info(f"Output saved: {output_file}")
            except Exception as e:
                logger.error(f"Error saving: {e}")
                return False
        
        # Archive
        try:
            archive_path = os.path.join(ARCHIVE_FOLDER, os.path.basename(input_file))
            shutil.move(input_file, archive_path)
            logger.info(f"Archived to: {archive_path}")
        except Exception as e:
            logger.error(f"Error archiving: {e}")
            return False
        
        logger.info(f"✓ Done: {input_file}\n")
        return True
        
    except Exception as e:
        logger.error(f"Error: {e}", exc_info=True)
        return False


def watch_folder():
    """Monitor upload folder and process files"""
    logger.info("Watcher started - monitoring upload/ folder")
    logger.info("Drop Excel files in upload/ - they will auto-process")
    logger.info("Press Ctrl+C to stop\n")
    
    processed = set()
    
    while True:
        try:
            # Check for files in upload folder
            files = list(Path(UPLOAD_FOLDER).glob("*"))
            
            for file_path in files:
                if file_path.is_file():
                    _, ext = os.path.splitext(file_path)
                    
                    if ext.lower() in SUPPORTED_EXTENSIONS and str(file_path) not in processed:
                        # Wait a moment for file to finish writing
                        time.sleep(1)
                        
                        if file_path.exists():  # Check if still there
                            processed.add(str(file_path))
                            process_file(str(file_path))
            
            time.sleep(2)  # Check every 2 seconds
            
        except KeyboardInterrupt:
            logger.info("Watcher stopped")
            break
        except Exception as e:
            logger.error(f"Error in watcher: {e}")
            time.sleep(2)


if __name__ == "__main__":
    watch_folder()
